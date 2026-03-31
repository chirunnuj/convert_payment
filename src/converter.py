import logging
import os
from datetime import datetime
from typing import List

import openpyxl

from src import constants
from src.models import ConversionResult, PaymentRecord
from src.validators import (
    ValidationError,
    format_amount,
    get_bank_info,
    get_bank_info_for_tcb,
    validate_account,
    validate_amount,
    validate_bank_name,
    validate_customer_name,
    validate_excel_headers,
    validate_payment_date,
    validate_payment_time,
)

logger = logging.getLogger(__name__)


def read_excel(input_path: str) -> List[PaymentRecord]:
    logger.info(f"Reading Excel file: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    all_headers = [cell.value for cell in ws[1]]
    headers = [h for h in all_headers if h is not None and str(h).strip() != ""]
    validate_excel_headers(headers)
    header_map = {h: idx for idx, h in enumerate(all_headers) if h is not None and str(h).strip() != ""}
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue
        try:
            bank_name = str(row[header_map["BankName"]]).strip()
            payment_date = str(row[header_map["PaymentDate"]]).strip()
            payment_time = str(row[header_map["PaymentTime"]]).strip()
            customer_name = str(row[header_map["CustomerName"]]).strip()
            account = str(row[header_map["Account"]]).strip()
            amount_str = str(row[header_map["Amt"]]).strip()
            
            validated_bank_name = validate_bank_name(bank_name)

            bank_info = get_bank_info(validated_bank_name)

            validate_payment_date(payment_date)
            validate_payment_time(payment_time)
            validate_customer_name(customer_name)
            validate_account(account)
            amount = validate_amount(amount_str)
            amount_formatted = format_amount(amount)
            record = PaymentRecord(
                bank_name=validated_bank_name,
                bank_code=bank_info["code"],
                company_account=bank_info["account"],
                payment_date=payment_date,
                payment_time=payment_time,
                customer_name=customer_name,
                account=account,
                amount=amount,
                amount_formatted=amount_formatted,
            )
            records.append(record)
        except ValidationError as e:
            logger.error(f"Validation error at row {row_idx}: {e}")
            raise
    logger.info(f"Read {len(records)} payment records")
    return records


def build_header_line(seq_no: int, effective_date: str, bank_code: str) -> str:
    line = [" "] * constants.LINE_LENGTH
    line[0] = "H"
    line[1:7] = list(str(seq_no).zfill(6))

    # bank_info = get_bank_info_for_tcb(constants.DEFAULT_BANK_NAME)
    line[7:10] = list(constants.DEFAULT_BANK_CODE)

    line[10:20] = list(constants.DEFAULT_COMPANY_ACCOUNT)
    line[20:60] = list(constants.COMPANY_NAME.ljust(40))
    line[60:68] = list(effective_date)
    return "".join(line)


def build_body_line(seq_no: int, record: PaymentRecord) -> str:
    line = [" "] * constants.LINE_LENGTH
    line[0] = "D"
    line[1:7] = list(str(seq_no).zfill(6))
    # logger.info(f"Bank Code: {record.bank_code}")
    bank_detail = constants.BANK_CODES_FOR_CUSTOMER[record.bank_name]
    # logger.info(f"Bank Detail: {bank_detail}")
    # logger.info(f"Bank Code: {bank_detail['code']}")
    # logger.info(f"Bank Account: {bank_detail['account']}")
    line[7:10] = list(bank_detail["code"])
    line[10:20] = list(bank_detail["account"])
    line[20:28] = list(record.payment_date)
    line[28:34] = list(record.payment_time)
    line[34:84] = list(record.customer_name.ljust(50))
    line[84:104] = list(record.account.ljust(20))
    line[104:124] = list(" " * 20)
    line[124:144] = list(" " * 20)
    line[144:148] = list("0901")
    line[148:152] = list("0000")
    line[152] = "C"
    line[153:156] = list("CSH")
    line[156:163] = list("0000000")
    line[163:176] = list(record.amount_formatted)
    return "".join(line)


def build_trailer_line(total_lines: int, bank_code: str, sum_credit: str) -> str:
    line = [" "] * constants.LINE_LENGTH
    line[0] = "T"
    line[1:7] = list(str(total_lines).zfill(6))
    line[7:10] = list(bank_code)
    line[7:10] = list(constants.DEFAULT_BANK_CODE)
    line[10:20] = list(constants.DEFAULT_COMPANY_ACCOUNT)
    line[20:33] = list("0000000000000")
    line[33:39] = list("000000")
    line[39:52] = list(sum_credit)

    logger.info(f"total_lines: {total_lines}")
    logger.info(f"body_lines: {total_lines-2}")

    line[52:58] = list(str(total_lines-2).zfill(6))
    # line[176:179] = list("000")
    return "".join(line)


def convert_excel_to_payment(input_path: str, output_dir: str = "output") -> ConversionResult:
    try:
        records = read_excel(input_path)
        if not records:
            return ConversionResult(
                success=False,
                error_message="No payment records found in input file",
            )
        effective_date = datetime.now().strftime("%d%m%Y")
        bank_code = records[0].bank_code
        header_line = build_header_line(1, effective_date, bank_code)
        body_lines = []
        total_amount = 0.0
        for idx, record in enumerate(records, start=2):
            body_line = build_body_line(idx, record)
            body_lines.append(body_line)
            total_amount += record.amount
        sum_credit = format_amount(total_amount)
        
        logger.info(f"total_amount: {total_amount}")
        logger.info(f"sum_credit: {sum_credit}")
        
        total_lines = len(records) + 2
        trailer_line = build_trailer_line(total_lines, bank_code, sum_credit)
        output_filename = f"COLLECTION_{effective_date}.txt"
        output_path = os.path.join(output_dir, output_filename)
        os.makedirs(output_dir, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(header_line + "\n")
            for body_line in body_lines:
                f.write(body_line + "\n")
            f.write(trailer_line + "\n")
        logger.info(f"Output written to: {output_path}")
        return ConversionResult(
            success=True,
            output_path=output_path,
            record_count=len(records),
            total_amount=total_amount,
        )
    except ValidationError as e:
        logger.error(f"Validation error: {e}")
        return ConversionResult(success=False, error_message=str(e))
    except Exception as e:
        logger.error(f"Conversion error: {e}")
        return ConversionResult(success=False, error_message=str(e))
